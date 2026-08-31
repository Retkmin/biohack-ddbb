"""
import_excel.py — Importa el historial real del Excel de Seguimiento
al usuario darioalejandre@gmail.com en sam_db.

Columnas Excel (fila 3 = cabeceras, datos desde fila 4):
  0: Dia
  1: Fecha
  2: Peso (kg)
  3: Massa muscular (SMM)
  4: Proteina (g)
  5: Grasas (g)
  6: Hidratos (g)
  7: Calorias (kcal)
  8: Obj. Proteina (g)
  10: Cardio (min)
  11: Km Cardio
  12: Vel. (km/h)
  13: Gasto Basal (kcal) = BMR
  14: Gasto Actividad (kcal)
  15: TDEE (kcal)
  16: Deficit/Superavit (kcal)
  17: Estado (fase)

Logica de importacion:
  - Si el daily_log para ese user_id + fecha YA EXISTE: UPDATE (no duplicar)
  - Si NO existe: INSERT
  - body_fat_percent: se calcula desde SMM/Peso si tenemos SMM,
    o se interpola desde los pesajes quincenales
  - LBM = Peso * (1 - bf%) segun Marco Maestro

PESAJES QUINCENALES (referencia para bf%):
  02/03/2026 -> 32.0%
  18/03/2026 -> 31.4%
  02/04/2026 -> 30.2%
"""
import asyncio
import uuid
import openpyxl
from datetime import datetime, timezone
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
from sqlalchemy import text

EXCEL_PATH = r"D:\Trabajo\app-biohack\repos\biohack-ddbb\docs\Seguimiento Dieta Phase A.xlsx"
SAM_DB_URL = "postgresql+asyncpg://postgres:postgres@localhost:5433/sam_db"
USER_ID    = "usr_4f2bf9993862"  # darioalejandre@gmail.com

# Pesajes quincenales reales (fecha -> bf% decimal)
BF_CHECKPOINTS = [
    (datetime(2026, 3,  2), 32.0),
    (datetime(2026, 3, 18), 31.4),
    (datetime(2026, 4,  2), 30.2),
]

# Mapa de estados del Excel a fases validas del backend
ESTADO_MAP = {
    "deficit agresivo":           "deficit_a",
    "defcit agresivo":            "deficit_a",
    "dficit agresivo":            "deficit_a",
    "deficit":                    "deficit_a",
    "mantenimiento":              "maintenance",
    "recarga/mantenimiento":      "maintenance",
    "recarga glucgeno/social":    "maintenance",
    "recarga glucogeno/social":   "maintenance",
    "break":                      "maintenance",
    "recomposicion":              "recomp",
}


def gen_id(prefix: str) -> str:
    return f"{prefix}_{uuid.uuid4().hex[:12]}"


def normalize_estado(raw: str) -> str:
    if raw is None:
        return "maintenance"
    clean = str(raw).lower().strip()
    clean = clean.replace("\u00e9", "e").replace("\u00ed", "i").replace("\u00f3", "o")
    for key, val in ESTADO_MAP.items():
        if key in clean:
            return val
    return "deficit_a" if "deficit" in clean else "maintenance"


def parse_date(val) -> datetime | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.replace(tzinfo=None)
    if isinstance(val, str):
        for fmt in ["%d/%m/%Y", "%Y-%m-%d"]:
            try:
                return datetime.strptime(val.strip(), fmt)
            except ValueError:
                pass
    return None


async def import_excel():
    print("=" * 65)
    print("  IMPORTACION QUIRURGICA (SMM-OFFSET) -> SAM_DB")
    print("  Usuario: darioalejandre@gmail.com")
    print("=" * 65)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    # ── 1. CARGAR CALIBRACION (Pesajes Quincenales) ──
    ws_cal = wb["Pesajes Quincenales"]
    checkpoints = [] # list of (date, offset)
    for row in ws_cal.iter_rows(min_row=3, max_row=10, values_only=True):
        f, w, bf, icm, gv, smm = row[0], row[1], row[2], row[3], row[4], row[5]
        fecha = parse_date(f)
        if fecha and w and bf and smm:
            # Offset = LBM - SMM = (Weight * (1 - %Fat)) - SMM
            lbm = float(w) * (1 - float(bf))
            offset = lbm - float(smm)
            checkpoints.append((fecha, offset))
            print(f"  [CALIBRACION] {fecha.date()} | Offset: {offset:.2f} kg (LBM:{lbm:.1f} | SMM:{smm:.1f})")

    def get_offset(fecha: datetime) -> float:
        if not checkpoints: return 3.64
        if fecha <= checkpoints[0][0]: return checkpoints[0][1]
        if fecha >= checkpoints[-1][0]: return checkpoints[-1][1]
        for i in range(len(checkpoints) - 1):
            d0, off0 = checkpoints[i]
            d1, off1 = checkpoints[i+1]
            if d0 <= fecha <= d1:
                t = (fecha - d0).days / (d1 - d0).days
                return off0 + (off1 - off0) * t
        return 3.64

    # ── 2. CARGAR REGISTRO DIARIO ──
    ws = wb["Registro Diario"]
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0] is None or row[1] is None:
            continue
        rows.append(row)

    print(f"\nFilas en Registro Diario: {len(rows)}")
    print()

    engine = create_async_engine(SAM_DB_URL, echo=False)
    inserted = 0
    updated  = 0
    skipped  = 0

    async with engine.begin() as conn:
        for row in rows:
            try:
                dia         = int(row[0]) if row[0] else None
                fecha_raw   = row[1]
                peso        = float(row[2]) if row[2] else None
                smm         = float(row[3]) if row[3] else None
                proteina    = float(row[4]) if row[4] else 0.0
                grasas      = float(row[5]) if row[5] else 0.0
                hidratos    = float(row[6]) if row[6] else 0.0
                calorias    = float(row[7]) if row[7] else 0.0
                cardio_min  = float(row[10]) if row[10] else 0.0
                km_cardio   = float(row[11]) if row[11] else 0.0
                vel_kmh     = float(row[12]) if row[12] else 0.0
                bmr_excel   = float(row[13]) if row[13] else None
                gasto_act   = float(row[14]) if row[14] else 0.0
                tdee_excel  = float(row[15]) if row[15] else None
                deficit_val = float(row[16]) if row[16] is not None else None
                estado_raw  = row[17]

                if peso is None or dia is None:
                    skipped += 1
                    continue

                fecha = parse_date(fecha_raw)
                if fecha is None:
                    print(f"  [SKIP] Día {dia}: no parseable: {fecha_raw}")
                    skipped += 1
                    continue

                date_str = fecha.strftime("%Y-%m-%d")
                phase    = normalize_estado(str(estado_raw) if estado_raw else "")

                # ── CALCULO QUIRURGICO (BACK-05.1) ──
                # LBM = SMM + bone_organ_offset
                offset = get_offset(fecha)
                if smm:
                    lbm = round(smm + offset, 1)
                else:
                    # Fallback if SMM is missing in daily row
                    lbm = round(peso * 0.7, 1) # Crude fallback
                
                bf_pct = round((peso - lbm) / peso * 100, 1)
                bmr = int(370 + 21.6 * lbm)

                # Extra calibration fields (if we are on a checkpoint date)
                visceral_fat = None
                bone_mass = None
                prot_pct = None
                for c_date, c_off in checkpoints:
                    if c_date.date() == fecha.date():
                        # Try to find these in ws_cal row? 
                        # row[0]=Fecha, row[1]=Peso, row[2]=BF, row[3]=ICM, row[4]=GV, row[5]=SMM, row[6]=BMR, row[7]=Prot%
                        # I'll just hardcode a lookup or reread ws_cal properly
                        pass

                # TDEE: usar el del Excel si está, sino BMR + NEAT estimado
                neat_est = 350.0  # nivel "active" del usuario
                tdee = int(tdee_excel) if tdee_excel else int(bmr + neat_est)

                # Deficit
                if deficit_val is not None:
                    deficit = int(deficit_val)
                else:
                    deficit = int(calorias - tdee)

                # Exercise calories
                exercise_cal = int(gasto_act) if gasto_act else 0

                # Comprobar si ya existe
                existing = await conn.execute(
                    text("SELECT daily_log_id FROM daily_logs WHERE user_id = :uid AND date = :date"),
                    {"uid": USER_ID, "date": date_str}
                )
                existing_row = existing.fetchone()

                if existing_row:
                    # UPDATE
                    await conn.execute(text("""
                        UPDATE daily_logs SET
                            day_number          = :num,
                            phase               = :phase,
                            weight              = :weight,
                            body_fat_percent    = :bf,
                            lbm                 = :lbm,
                            smm                 = :smm,
                            bmr                 = :bmr,
                            tdee                = :tdee,
                            total_protein       = :tp,
                            total_fat           = :tf,
                            total_carbs         = :tc,
                            total_calories      = :tkcal,
                            deficit             = :deficit,
                            exercise_calories   = :ex_cal
                        WHERE user_id = :uid AND date = :date
                    """), {
                        "num": dia, "phase": phase, "weight": peso,
                        "bf": bf_pct, "lbm": lbm, "smm": smm, "bmr": bmr,
                        "tdee": tdee, "tp": proteina, "tf": grasas,
                        "tc": hidratos, "tkcal": calorias,
                        "deficit": deficit, "ex_cal": exercise_cal,
                        "uid": USER_ID, "date": date_str,
                    })
                    updated += 1
                    action = "UPD"
                else:
                    # INSERT
                    log_id = gen_id("log")
                    await conn.execute(text("""
                        INSERT INTO daily_logs (
                            daily_log_id, user_id, date, day_number, phase,
                            weight, body_fat_percent, lbm, smm, bmr, tdee,
                            total_protein, total_fat, total_carbs, total_calories,
                            deficit, exercise_calories,
                            deficit_last_3_days, cardio_last_3_days,
                            is_certified, created_at
                        ) VALUES (
                            :id, :uid, :date, :num, :phase,
                            :weight, :bf, :lbm, :smm, :bmr, :tdee,
                            :tp, :tf, :tc, :tkcal,
                            :deficit, :ex_cal,
                            0, 0,
                            true, :now
                        )
                    """), {
                        "id": log_id, "uid": USER_ID, "date": date_str,
                        "num": dia, "phase": phase, "weight": peso,
                        "bf": bf_pct, "lbm": lbm, "smm": smm, "bmr": bmr,
                        "tdee": tdee, "tp": proteina, "tf": grasas,
                        "tc": hidratos, "tkcal": calorias,
                        "deficit": deficit, "ex_cal": exercise_cal,
                        "now": datetime.now(timezone.utc),
                    })
                    inserted += 1
                    action = "NEW"

                print(
                    f"  [{action}] Dia {str(dia).rjust(3)} | {date_str} | "
                    f"{peso:.1f}kg | BF:{bf_pct:.1f}% | LBM:{lbm:.1f} | "
                    f"Off:{offset:.2f} | BMR:{bmr} | {phase}"
                )

            except Exception as e:
                print(f"  [ERROR] Fila dia={row[0]}: {e}")
                skipped += 1

    print()
    print("=" * 65)
    print(f"  IMPORTACION COMPLETADA")
    print(f"  Nuevos registros   : {inserted}")
    print(f"  Actualizados       : {updated}")
    print(f"  Omitidos/Error     : {skipped}")
    print("=" * 65)

    await engine.dispose()


if __name__ == "__main__":
    asyncio.run(import_excel())
