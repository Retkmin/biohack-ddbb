"""
inspect_excel.py - Inspecciona la estructura del Excel de Seguimiento.
"""
import sys
import openpyxl

EXCEL_PATH = r"D:\Trabajo\app-biohack\repos\biohack-ddbb\docs\Seguimiento Dieta Phase A.xlsx"

def inspect():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print("Hojas: " + str(wb.sheetnames))

    # Hoja principal de datos
    ws = wb["Registro Diario"]
    print()
    print("=== HOJA: Registro Diario | Filas: " + str(ws.max_row) + " ===")
    print()

    # Cabeceras (fila 3)
    headers = [str(cell.value) if cell.value is not None else "" for cell in ws[3]]
    print("CABECERAS:")
    for i, h in enumerate(headers):
        if h:
            print("  Col " + str(i) + ": " + h)

    print()
    print("DATOS (filas 4 en adelante):")
    print("-" * 110)
    header_line = " | ".join([
        "Dia".ljust(4), "Fecha".ljust(12), "Peso".ljust(7), "SMM".ljust(7),
        "Prot".ljust(6), "Gras".ljust(6), "Hidr".ljust(6), "Kcal".ljust(6),
        "ObjP".ljust(6), "Cardio(m)".ljust(10), "Km".ljust(6), "Vel".ljust(5),
        "GastoB".ljust(8), "GastoA".ljust(8), "TDEE".ljust(6), "Deficit".ljust(8), "Estado"
    ])
    print(header_line)
    print("-" * 110)

    rows_with_data = 0
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0] is None:
            continue
        rows_with_data += 1
        dia     = str(row[0]).split(".")[0] if row[0] else ""
        fecha   = str(row[1]) if row[1] else ""
        peso    = str(row[2]) if row[2] else ""
        smm     = str(round(row[3], 2)) if row[3] else ""
        prot    = str(row[4]) if row[4] else ""
        gras    = str(row[5]) if row[5] else ""
        hidr    = str(row[6]) if row[6] else ""
        kcal    = str(row[7]) if row[7] else ""
        obj_p   = str(row[8]) if row[8] else ""
        cardio  = str(row[10]) if row[10] else ""
        km      = str(row[11]) if row[11] else ""
        vel     = str(row[12]) if row[12] else ""
        gasto_b = str(row[13]) if row[13] else ""
        gasto_a = str(row[14]) if row[14] else ""
        tdee    = str(row[15]) if row[15] else ""
        deficit = str(row[16]) if row[16] else ""
        estado  = str(row[17]) if row[17] else ""

        line = " | ".join([
            dia.ljust(4), fecha.ljust(12), peso.ljust(7), smm.ljust(7),
            prot.ljust(6), gras.ljust(6), hidr.ljust(6), kcal.ljust(6),
            obj_p.ljust(6), cardio.ljust(10), km.ljust(6), vel.ljust(5),
            gasto_b.ljust(8), gasto_a.ljust(8), tdee.ljust(6), deficit.ljust(8), estado
        ])
        print(line)

    print("-" * 110)
    print("Total filas con datos: " + str(rows_with_data))

    # Hoja de pesajes quincenales
    print()
    print("=== HOJA: Pesajes Quincenales ===")
    ws2 = wb["Pesajes Quincenales"]
    for row in ws2.iter_rows(min_row=1, max_row=15, values_only=True):
        if any(v is not None for v in row):
            print("  " + str(row))

    # Configuracion
    print()
    print("=== HOJA: Configuracion ===")
    ws3 = wb["Configuracion"]
    for row in ws3.iter_rows(min_row=1, max_row=20, values_only=True):
        if any(v is not None for v in row):
            print("  " + str(row))


inspect()
